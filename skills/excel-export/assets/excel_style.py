"""
excel-export skill helpers — domain-agnostic openpyxl building blocks for the
house style (navy/red, gridlines off, live delta formulas, charts sized to the
content). Applies to any analysis, not just the energy example the style was
reverse-engineered from.

The workhorse is `bs_table`: it writes a whole basis/scenarie/Δ table (header,
rows, live Δ formulas, optional SUM total) from a column spec + a list of rows,
so a typical comparison table is ~5 lines, not a hand-written loop. Reach for the
low-level helpers (`data_row`, `delta_formula`, direct cell writes) only for the
odd table `bs_table` can't express (section-grouped totals, decompositions).

`finalize` recalculates the file and reads back your check cells IN ONE process,
so the build script self-verifies with no extra shell round-trip.
"""
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import get_column_letter

# ---- palette --------------------------------------------------------------
NAVY = "1F2A44"        # primary text, header fill, primary series
SUBHEADER = "41546E"   # lighter navy, sub-header fill (paired-group columns)
ACCENT = "E04131"      # subtitles, delta values, secondary series
LINK_GREEN = "008000"  # cross-sheet reference (provenance marker only)
ZEBRA = "EEF1F5"       # aggregate / total row background
LABEL_GRAY = "7A8699"  # small section dividers ("Aggregater", "I alt")
GRID_GRAY = "878787"   # chart gridlines
WHITE = "FFFFFF"

# ---- number formats -------------------------------------------------------
# Pick by magnitude/nature of the value, not by habit — see SKILL.md "Number formats".
FMT_3DP = "0.000"                              # precise small-magnitude levels
FMT_DELTA_3DP = r"\+0.000;\-0.000;0"           # deltas on that scale
FMT_2DP = "0.00"                               # two-decimal level (e.g. TWh)
FMT_DELTA_2DP = r"\+0.00;\-0.00;0.00"          # signed two-decimal delta
FMT_2DP_PAREN = r"0.00;\(0.00\)"               # 2dp, negative = reduction, in parens
FMT_1DP = "0.0"                                # small level, one decimal (prices, rates)
FMT_DELTA_1DP = r"\+0.0;\-0.0;0.0"             # signed one-decimal delta
FMT_INT = "#,##0"                              # large levels / counts
FMT_DELTA_INT = r"\+#,##0;\-#,##0;0"           # signed integer deltas
FMT_INT_PAREN = r"#,##0;\(#,##0\)"             # large value, negative = reduction, in parens
FMT_MDKK_1DP = r"#,##0.0;\(#,##0.0\)"          # large value, 1dp, negative in parens (mDKK)
FMT_DELTA_MDKK_1DP = r"\+#,##0.0;\-#,##0.0;0.0"  # signed one-decimal large-value delta
FMT_PCT1 = "0.0%"                              # shares / rates
FMT_DELTA_PCT1 = r"\+0.0%;\-0.0%;0.0%"         # signed percentage-point deltas


# ---- sheet / text scaffolding --------------------------------------------
def setup_sheet(ws, zoom=130, margin_width=3.0):
    """Gridlines off, house zoom level, narrow column A as left margin."""
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    ws.column_dimensions["A"].width = margin_width


def write_title(ws, title, subtitle="", row=2):
    """B{row}: bold 14pt navy headline. B{row+1}: 10pt red subtitle (units/scope)."""
    c = ws.cell(row=row, column=2, value=title)
    c.font = Font(bold=True, size=14, color=NAVY)
    ws.row_dimensions[row].height = 17.55
    if subtitle:
        ws.cell(row=row + 1, column=2, value=subtitle).font = Font(size=10, color=ACCENT)


def header_row(ws, row, headers, start_col=2):
    """Bold white text on navy fill, centered — top-level column headers."""
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")


def subheader_row(ws, row, labels, start_col=3):
    """Bold white on lighter navy — second header row for base/shock/Δ/Δ% groups."""
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=lab)
        c.font = Font(bold=True, size=9, color=WHITE)
        c.fill = PatternFill("solid", fgColor=SUBHEADER)
        c.alignment = Alignment(horizontal="center")


def section_label(ws, row, text, col=2):
    """Small gray divider above a sub-block, e.g. 'Aggregater (vægtet)'."""
    ws.cell(row=row, column=col, value=text).font = Font(bold=True, size=9, color=LABEL_GRAY)


def data_row(ws, row, label, values, start_col=3, numfmt=FMT_INT, bold_label=True, zebra=False):
    """One uniform data row: bold navy label in col B, centered values from start_col.
    For rows needing per-cell formats / a gloss column / a live Δ, prefer `bs_table`
    or write the cells directly."""
    lc = ws.cell(row=row, column=2, value=label)
    lc.font = Font(bold=bold_label, size=10, color=ACCENT if zebra else NAVY)
    if zebra:
        lc.fill = PatternFill("solid", fgColor=ZEBRA)
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.font = Font(bold=bold_label, size=10, color=NAVY)
        c.number_format = numfmt
        c.alignment = Alignment(horizontal="center")
        if zebra:
            c.fill = PatternFill("solid", fgColor=ZEBRA)


def delta_formula(ws, row, col, minuend_col, subtrahend_col, numfmt=FMT_DELTA_INT):
    """Write a live =B-A formula (never a hardcoded number) styled as a delta: bold, red."""
    col_l = get_column_letter(col)
    m_l = get_column_letter(minuend_col)
    s_l = get_column_letter(subtrahend_col)
    c = ws.cell(row=row, column=col, value=f"={m_l}{row}-{s_l}{row}")
    c.font = Font(bold=True, size=10, color=ACCENT)
    c.number_format = numfmt
    c.alignment = Alignment(horizontal="center")
    return c


def _valcell(ws, r, c, value, numfmt, *, bold=False, zebra=False, color=NAVY):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = Font(bold=bold, size=10, color=color)
    cell.number_format = numfmt
    cell.alignment = Alignment(horizontal="center")
    if zebra:
        cell.fill = PatternFill("solid", fgColor=ZEBRA)
    return cell


# ---- the workhorse: a whole basis/scenarie/Δ table from a spec ------------
def bs_table(ws, start_row, columns, rows, *, title=None, total=None,
             start_col=2, freeze=False, gap_before_total=1):
    """Write a base/scenarie/Δ table from a column spec + rows. Returns row anchors.

    columns: list of dicts, left→right from `start_col`. Each is one of:
      {"h": "Zone",     "kind": "label"}
      {"h": "Basis",    "kind": "val",   "fmt": FMT_1DP}
      {"h": "Scenarie", "kind": "val",   "fmt": FMT_1DP}
      {"h": "Δ",        "kind": "delta", "fmt": FMT_DELTA_1DP, "of": ("Scenarie", "Basis")}
    A "delta" column writes a live formula = <first> − <second> from `of`.
    A plain extra text column (e.g. a Danish gloss) is just a {"kind": "val"} with a
    text value and any fmt (format is ignored for strings).

    rows: list of tuples — one value per NON-delta column, in column order (label
    first, then each val). Delta columns are computed, not supplied.

    total: None, or a label string → a zebra total row: label = the string,
    each val col = SUM over the data rows, each delta col = the same Δ formula.

    Returns {"header", "first", "last", "total"(or None), "cols": {title: col_idx}}.
    Set freeze=True on the first/only table of a sheet to freeze under its header.
    """
    r_head = start_row
    if title is not None:
        section_label(ws, start_row, title, col=start_col)
        r_head = start_row + 1
    header_row(ws, r_head, [c["h"] for c in columns], start_col=start_col)
    cols = {c["h"]: start_col + i for i, c in enumerate(columns)}
    input_cols = [c for c in columns if c["kind"] != "delta"]

    r0 = r_head + 1
    r = r0
    for row in rows:
        if len(row) != len(input_cols):
            raise ValueError(f"row {row!r}: expected {len(input_cols)} values, got {len(row)}")
        vi = 0
        for c in columns:
            ci = cols[c["h"]]
            if c["kind"] == "label":
                lc = ws.cell(row=r, column=ci, value=row[vi]); vi += 1
                lc.font = Font(bold=True, size=10, color=NAVY)
            elif c["kind"] == "val":
                v = row[vi]; vi += 1
                _valcell(ws, r, ci, v, c.get("fmt", FMT_INT))
            elif c["kind"] == "delta":
                plus, minus = c["of"]
                delta_formula(ws, r, ci, cols[plus], cols[minus], numfmt=c.get("fmt", FMT_DELTA_INT))
            else:
                raise ValueError(f"unknown column kind {c['kind']!r}")
        r += 1
    r_last = r - 1

    r_total = None
    if total is not None:
        r_total = r_last + 1 + gap_before_total
        for c in columns:
            ci = cols[c["h"]]
            if c["kind"] == "label":
                lc = ws.cell(row=r_total, column=ci, value=total)
                lc.font = Font(bold=True, size=10, color=ACCENT)
                lc.fill = PatternFill("solid", fgColor=ZEBRA)
            elif c["kind"] == "val":
                cl = get_column_letter(ci)
                cell = ws.cell(row=r_total, column=ci, value=f"=SUM({cl}{r0}:{cl}{r_last})")
                cell.font = Font(bold=True, size=10, color=NAVY)
                cell.number_format = c.get("fmt", FMT_INT)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
            elif c["kind"] == "delta":
                plus, minus = c["of"]
                dcell = delta_formula(ws, r_total, ci, cols[plus], cols[minus],
                                      numfmt=c.get("fmt", FMT_DELTA_INT))
                dcell.fill = PatternFill("solid", fgColor=ZEBRA)

    if freeze:
        ws.freeze_panes = ws.cell(row=r0, column=start_col).coordinate
    return {"header": r_head, "first": r0, "last": r_last, "total": r_total, "cols": cols}


# ---- overview cross-sheet links (green = provenance) ----------------------
def link_cell(ws, row, col, ref, numfmt, *, bold=False, zebra=False):
    """A single GREEN cross-sheet reference cell. `ref` is a full ref WITH quoting
    if the sheet name has a space/period, e.g. "'3. Priser'!C9"."""
    cell = ws.cell(row=row, column=col, value=f"={ref}")
    cell.font = Font(bold=bold, size=10, color=LINK_GREEN)
    cell.number_format = numfmt
    cell.alignment = Alignment(horizontal="center")
    if zebra:
        cell.fill = PatternFill("solid", fgColor=ZEBRA)
    return cell


def linkrow(ws, row, label, ref_base, ref_shock, lvl_fmt, d_fmt, *, start_col=2):
    """Overview nøgletal row: basis + scenarie as GREEN cross-sheet refs, Δ as a
    local red formula on top of them. refs are full & quoted, e.g. "'2. Priser'!C8"."""
    lc = ws.cell(row=row, column=start_col, value=label)
    lc.font = Font(size=10, color=NAVY)
    link_cell(ws, row, start_col + 1, ref_base, lvl_fmt)
    link_cell(ws, row, start_col + 2, ref_shock, lvl_fmt)
    delta_formula(ws, row, start_col + 3, start_col + 2, start_col + 1, numfmt=d_fmt)


# ---- charts ---------------------------------------------------------------
def _finish_chart(ws, chart, anchor, title, cats_ref, data_ref, colors,
                   legend, numfmt, width, height):
    """Shared house styling: gray horizontal gridlines, navy/red series, sized legend."""
    chart.title = title
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=GRID_GRAY))
    chart.y_axis.numFmt = numfmt
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    for s, color in zip(chart.series, colors):
        s.graphicalProperties.solidFill = color
        if isinstance(chart, LineChart):
            s.graphicalProperties.line.solidFill = color
            s.graphicalProperties.line.width = 20000  # EMU, ~1.6pt
    if legend:
        chart.legend.position = legend
    else:
        chart.legend = None
    chart.width = width
    chart.height = height
    ws.add_chart(chart, anchor)
    return chart


def add_bar_chart(ws, anchor, title, cats_ref, data_ref, colors=(NAVY, ACCENT),
                   legend="r", numfmt=FMT_INT, width=15, height=7.5, gap_width=60):
    """Clustered column chart — the default for cross-sectional comparisons across
    categories. legend="r" for 2+ series; legend=None + colors=(NAVY,) for one series."""
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.gapWidth = gap_width
    return _finish_chart(ws, chart, anchor, title, cats_ref, data_ref, colors,
                          legend, numfmt, width, height)


def add_line_chart(ws, anchor, title, cats_ref, data_ref, colors=(NAVY, ACCENT),
                    legend="r", numfmt=FMT_INT, width=15, height=7.5):
    """Line chart — for a time series (year-by-year, period-by-period)."""
    chart = LineChart()
    return _finish_chart(ws, chart, anchor, title, cats_ref, data_ref, colors,
                          legend, numfmt, width, height)


# ---- one-shot recalc + verify (no extra shell round-trip) -----------------
def finalize(path, checks=None, *, recalc_script="/mnt/skills/public/xlsx/scripts/recalc.py",
             show=True):
    """Recalc the workbook via LibreOffice and read back `checks` — in ONE process.
    Call this at the END of the build script; it replaces the separate recalc/verify
    shell steps. Render a page image ONLY if this reports errors (or the user asks).

    checks: dict {label: "Sheet!Cell"} — sheet name may be quoted or not.
    Returns (total_errors, {label: value}). Prints a compact report when show=True.
    """
    import subprocess, sys, os, json
    from openpyxl import load_workbook

    path = os.path.abspath(path)
    proc = subprocess.run(
        [sys.executable, os.path.basename(recalc_script), path],
        cwd=os.path.dirname(recalc_script), capture_output=True, text=True,
    )
    info, errs = {}, None
    try:
        out = proc.stdout.strip()
        info = json.loads(out[out.index("{"):out.rindex("}") + 1])
        errs = info.get("total_errors")
    except Exception:
        if show:
            print("recalc: could not parse output:\n", proc.stdout[-500:], proc.stderr[-300:])
    if show:
        print(f"recalc: total_errors={errs}  formulas={info.get('total_formulas', '?')}")
        if errs:
            print("  error_summary:", info.get("error_summary"))

    values = {}
    if checks:
        wb = load_workbook(path, data_only=True)
        for name, addr in checks.items():
            sheet, cell = addr.rsplit("!", 1)
            v = wb[sheet.strip().strip("'")][cell].value
            values[name] = v
            if show:
                vs = f"{v:,.3f}" if isinstance(v, (int, float)) else v
                print(f"  {name:34s} {addr:26s} = {vs}")
    return errs, values